import os
from datetime import date

import pytest
from django.urls import reverse

from api.models import (
    GFCRFinanceSolution,
    GFCRIndicatorSet,
    GFCRRevenue,
    ProjectProfile,
    RestrictedProjectSummarySampleEvent,
)


@pytest.fixture
def create_indicator_set_payload():
    return {
        "title": "Dustin's IS",
        "report_date": "2024-02-10",
        "report_year": 2024,
        "indicator_set_type": "report",
        "f4_start_date": "1970-01-01",
        "f4_end_date": "2024-04-29",
        "finance_solutions": [
            {
                "name": "My FS",
                "sector": "ce_waste_management",
                "revenues": [{"revenue_type": "debt_conversion"}],
                "investment_sources": [
                    {"investment_source": "public", "investment_type": "concessional_loan"},
                    {"investment_source": "philanthropy", "investment_type": "financial_guarantee"},
                ],
            }
        ],
    }


@pytest.fixture
def indicator_set(project1):
    return GFCRIndicatorSet.objects.create(
        project=project1,
        title="Dustin's IS",
        report_date="2024-02-10",
        indicator_set_type="report",
        f4_start_date=date(1970, 1, 1),
        f4_end_date=date(2024, 4, 19),
    )


@pytest.fixture
def finance_solution(indicator_set):
    return GFCRFinanceSolution.objects.create(
        indicator_set=indicator_set,
        name="My FS",
        sector="ce_waste_management",
    )


@pytest.fixture
def revenue(finance_solution):
    return GFCRRevenue.objects.create(
        finance_solution=finance_solution,
        revenue_type="debt_conversion",
    )


@pytest.fixture
def restricted_project_summary_sample_events(
    project1,
):
    return RestrictedProjectSummarySampleEvent.objects.create(
        project_id=project1.id,
        records=[
            {
                "sample_date": "2024-01-01",
                "protocols": {
                    "beltfish": {
                        "biomass_kgha_avg": 0,
                    },
                    "benthiclit": {
                        "percent_cover_benthic_category_avg": {"Hard coral": 5, "Macroalgae": 10},
                    },
                    "benthicpit": {
                        "percent_cover_benthic_category_avg": {"Hard coral": 10, "Macroalgae": 20},
                    },
                    "benthicpqt": {
                        "percent_cover_benthic_category_avg": {"Hard coral": 15, "Macroalgae": 30},
                    },
                },
            },
            {
                "sample_date": "2024-01-02",
                "protocols": {
                    "beltfish": {
                        "biomass_kgha_avg": 10,
                    },
                    "benthiclit": {
                        "percent_cover_benthic_category_avg": {"Hard coral": 20, "Macroalgae": 40}
                    },
                    "benthicpit": {
                        "percent_cover_benthic_category_avg": {"Hard coral": 25, "Macroalgae": 50}
                    },
                    "benthicpqt": {
                        "percent_cover_benthic_category_avg": {"Hard coral": 30, "Macroalgae": 60}
                    },
                },
            },
        ],
    )


def test_create_indicator_set_admin(
    db_setup,
    api_client1,
    project1,
    project_profile1,
    create_indicator_set_payload,
):
    project1.includes_gfcr = True
    project1.save()

    url = reverse("indicatorset-list", kwargs={"project_pk": project1.pk})
    request = api_client1.post(url, data=create_indicator_set_payload, format="json")
    print(create_indicator_set_payload)

    assert request.status_code == 201
    response_data = request.json()

    indicator_set = GFCRIndicatorSet.objects.get(project=project1)
    finance_solution = indicator_set.finance_solutions.all()[0]
    investment_source = finance_solution.investment_sources.all()[0]
    revenue = finance_solution.revenues.all()[0]

    assert str(indicator_set.pk) == response_data["id"]

    finance_solution_data = response_data["finance_solutions"][0]

    assert finance_solution_data["id"] is not None
    assert str(finance_solution.pk) == finance_solution_data["id"]

    investment_source_data = finance_solution_data["investment_sources"][0]
    assert investment_source_data["id"] is not None
    assert str(investment_source.pk) == investment_source_data["id"]

    revenue_data = finance_solution_data["revenues"][0]
    assert revenue_data["id"] is not None
    assert str(revenue.pk) == revenue_data["id"]


def test_create_indicator_set_non_admin(
    db_setup,
    api_client1,
    project1,
    project_profile1,
    create_indicator_set_payload,
):
    project_profile1.role = ProjectProfile.COLLECTOR
    project_profile1.save()
    url = reverse("indicatorset-list", kwargs={"project_pk": project1.pk})
    request = api_client1.post(url, data=create_indicator_set_payload, format="json")

    assert request.status_code == 403


def test_update_indicator_set(
    db_setup,
    api_client1,
    project1,
    indicator_set,
):
    project1.includes_gfcr = True
    project1.save()

    # Create an indicator set
    url = reverse("indicatorset-detail", kwargs={"project_pk": project1.pk, "pk": indicator_set.id})
    create_request = api_client1.get(url, None, format="json")
    assert create_request.status_code == 200
    response_data = create_request.json()

    # Update the indicator set
    update_url = reverse(
        "indicatorset-detail", kwargs={"project_pk": project1.pk, "pk": indicator_set.id}
    )
    update_payload = response_data
    update_payload["title"] = "Updated Indicator Set"
    update_request = api_client1.put(update_url, data=update_payload, format="json")
    assert update_request.status_code == 200
    updated_response_data = update_request.json()

    # Verify the updated indicator set
    assert updated_response_data["title"] == "Updated Indicator Set"
    assert updated_response_data["id"] == str(indicator_set.id)


def test_coral_health_calculations(
    db_setup,
    api_client1,
    project1,
    indicator_set,
    restricted_project_summary_sample_events,
):
    project1.includes_gfcr = True
    project1.save()

    url = reverse("indicatorset-detail", kwargs={"project_pk": project1.pk, "pk": indicator_set.id})
    request = api_client1.get(url, None, format="json")

    assert request.status_code == 200
    response_data = request.json()

    assert response_data["f4_1_calc"] == 17.5
    assert response_data["f4_2_calc"] == 35
    assert response_data["f4_3_calc"] == 5


def test_reporting_range(
    db_setup,
    api_client1,
    project1,
    create_indicator_set_payload,
    belt_fish_project,
    update_summary_cache,
):
    project1.includes_gfcr = True
    project1.save()

    url = reverse("indicatorset-list", kwargs={"project_pk": project1.pk})
    request = api_client1.post(url, data=create_indicator_set_payload, format="json")
    assert request.status_code == 201
    response_data = request.json()
    assert response_data["f4_3"] is not None

    indicator_set = GFCRIndicatorSet.objects.get(pk=response_data["id"])
    indicator_set.f4_end_date = date(1980, 1, 1)
    indicator_set.save()

    url = reverse("indicatorset-detail", kwargs={"project_pk": project1.pk, "pk": indicator_set.id})
    request = api_client1.get(url, None, format="json")

    assert request.status_code == 200
    response_data = request.json()
    assert response_data["f4_3_calc"] is None


def test_notes_in_report_export(
    db_setup,
    project1,
):
    """Test that all notes fields are included in the GFCR report export."""
    from openpyxl import load_workbook

    from api.reports.gfcr import create_report

    project1.includes_gfcr = True
    project1.save()

    # Create indicator set with all notes fields populated
    indicator_set = GFCRIndicatorSet.objects.create(
        project=project1,
        title="Test Indicator Set",
        report_date="2024-02-10",
        indicator_set_type="report",
        f1_notes="F1 test notes",
        f2_notes="F2 test notes",
        f3_notes="F3 test notes",
        f4_notes="F4 test notes",
        f4_start_date=date(1970, 1, 1),
        f4_end_date=date(2024, 4, 19),
        f5_notes="F5 test notes",
        f6_notes="F6 test notes",
        f7_notes="F7 test notes",
    )

    # Create finance solution with notes
    finance_solution = GFCRFinanceSolution.objects.create(
        indicator_set=indicator_set,
        name="Test Finance Solution",
        sector="ce_waste_management",
        notes="Finance solution notes",
    )

    # Create investment with notes
    from api.models import GFCRInvestmentSource

    GFCRInvestmentSource.objects.create(
        finance_solution=finance_solution,
        investment_source="public",
        investment_type="grant",
        investment_amount=10000,
        notes="Investment notes",
    )

    # Create revenue with notes
    GFCRRevenue.objects.create(
        finance_solution=finance_solution,
        revenue_type="ecotourism",
        revenue_amount=5000,
        notes="Revenue notes",
    )

    # Generate report
    report_path = create_report([project1.id])
    assert report_path is not None

    # Load the workbook and verify notes are present
    wb = load_workbook(report_path)

    # Check F1 sheet - should have notes in column 8
    f1_sheet = wb["F1"]
    f1_row = list(f1_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert f1_row[7] == "F1 test notes", f"F1 notes not found. Row: {f1_row}"

    # Check F2 sheet - should have notes in column 7
    f2_sheet = wb["F2"]
    f2_rows = list(f2_sheet.iter_rows(min_row=2, values_only=True))
    # All F2 rows should have the same notes
    for row in f2_rows:
        if row[0] is not None:  # Skip empty rows
            assert row[6] == "F2 test notes", f"F2 notes not found in row: {row}"
            break

    # Check F3 sheet - should have notes in column 7
    f3_sheet = wb["F3"]
    f3_rows = list(f3_sheet.iter_rows(min_row=2, values_only=True))
    for row in f3_rows:
        if row[0] is not None:
            assert row[6] == "F3 test notes", f"F3 notes not found in row: {row}"
            break

    # Check F4 sheet - should have notes in column 9
    f4_sheet = wb["F4"]
    f4_rows = list(f4_sheet.iter_rows(min_row=2, values_only=True))
    for row in f4_rows:
        if row[0] is not None:
            assert row[8] == "F4 test notes", f"F4 notes not found in row: {row}"
            break

    # Check F5 sheet - should have notes in column 7
    f5_sheet = wb["F5"]
    f5_rows = list(f5_sheet.iter_rows(min_row=2, values_only=True))
    for row in f5_rows:
        if row[0] is not None:
            assert row[6] == "F5 test notes", f"F5 notes not found in row: {row}"
            break

    # Check F6 sheet - should have notes in column 7
    f6_sheet = wb["F6"]
    f6_rows = list(f6_sheet.iter_rows(min_row=2, values_only=True))
    for row in f6_rows:
        if row[0] is not None:
            assert row[6] == "F6 test notes", f"F6 notes not found in row: {row}"
            break

    # Check F7 sheet - should have notes in column 7
    f7_sheet = wb["F7"]
    f7_rows = list(f7_sheet.iter_rows(min_row=2, values_only=True))
    for row in f7_rows:
        if row[0] is not None:
            assert row[6] == "F7 test notes", f"F7 notes not found in row: {row}"
            break

    # Check FacilitiesSolutions sheet - notes are now at column 16 (index 15)
    bfs_sheet = wb["FacilitiesSolutions"]
    bfs_row = list(bfs_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert (
        bfs_row[15] == "Finance solution notes"
    ), f"Finance solution notes not found. Row: {bfs_row}"

    # Check Investments sheet - notes are now at column 15 (index 14)
    inv_sheet = wb["Investments"]
    inv_row = list(inv_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert inv_row[14] == "Investment notes", f"Investment notes not found. Row: {inv_row}"

    # Check Revenues sheet - notes are now at column 15 (index 14)
    rev_sheet = wb["Revenues"]
    rev_row = list(rev_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert rev_row[14] == "Revenue notes", f"Revenue notes not found. Row: {rev_row}"

    # Clean up
    os.remove(report_path)


def test_finance_solution_new_fields_in_serializer(
    db_setup,
    api_client1,
    project1,
    indicator_set,
):
    project1.includes_gfcr = True
    project1.save()

    GFCRFinanceSolution.objects.create(
        indicator_set=indicator_set,
        name="FS with new fields",
        fs_type="ctf",
        geographical_coverage="national",
        taf_name="",
        number_of_solutions_supported_by=0,
    )

    url = reverse("indicatorset-detail", kwargs={"project_pk": project1.pk, "pk": indicator_set.id})
    response = api_client1.get(url, format="json")
    assert response.status_code == 200
    fs_data = response.json()["finance_solutions"][0]
    assert fs_data["fs_type"] == "ctf"
    assert fs_data["geographical_coverage"] == "national"
    assert fs_data["taf_name"] == ""
    assert fs_data["number_of_solutions_supported_by"] == 0


def test_finance_solution_validate_passes_when_type_none(
    db_setup,
    api_client1,
    project1,
    project_profile1,
    indicator_set,
):
    project1.includes_gfcr = True
    project1.save()

    payload = {
        "id": str(indicator_set.id),
        "title": indicator_set.title,
        "report_date": str(indicator_set.report_date),
        "indicator_set_type": indicator_set.indicator_set_type,
        "f4_start_date": str(indicator_set.f4_start_date),
        "f4_end_date": str(indicator_set.f4_end_date),
        "project": str(project1.pk),
        "finance_solutions": [
            {
                "name": "FS no type",
                "sector": "ce_waste_management",
                "geographical_coverage": "national",
                "used_an_incubator": "gfcr_funded",
                "gender_smart": True,
                "sustainable_finance_mechanisms": ["blue_bonds"],
                "number_of_solutions_supported_by": 5,
                # fs_type intentionally omitted — validation must be skipped
            }
        ],
    }
    url = reverse("indicatorset-detail", kwargs={"project_pk": project1.pk, "pk": indicator_set.id})
    response = api_client1.put(url, data=payload, format="json")
    assert response.status_code == 200


def _fs_put_payload(indicator_set, project, fs_fields):
    return {
        "id": str(indicator_set.id),
        "title": indicator_set.title,
        "report_date": str(indicator_set.report_date),
        "indicator_set_type": indicator_set.indicator_set_type,
        "f4_start_date": str(indicator_set.f4_start_date),
        "f4_end_date": str(indicator_set.f4_end_date),
        "project": str(project.pk),
        "finance_solutions": [fs_fields],
    }


_FS_BASE = {
    "name": "FS",
    "fs_type": "business",
    "sector": "ce_waste_management",
    "geographical_coverage": "",
    "used_an_incubator": "",
    "taf_name": "",
    "number_of_solutions_supported_by": 0,
    "local_enterprise": False,
    "gender_smart": False,
    "sustainable_finance_mechanisms": [],
}


@pytest.mark.parametrize(
    "override,expected_error_field",
    [
        # sector is required for Business — no value to coerce to
        ({"fs_type": "business", "sector": ""}, "sector"),
        # geographical_coverage is required for CTF — no value to coerce to
        ({"fs_type": "ctf", "geographical_coverage": ""}, "geographical_coverage"),
        # number_of_solutions_supported_by must be > 0 for TAF — no value to coerce to
        (
            {"fs_type": "taf", "number_of_solutions_supported_by": 0},
            "number_of_solutions_supported_by",
        ),
    ],
)
def test_finance_solution_validate_errors(
    db_setup, api_client1, project1, project_profile1, indicator_set, override, expected_error_field
):
    project1.includes_gfcr = True
    project1.save()
    fs = {**_FS_BASE, **override}
    url = reverse("indicatorset-detail", kwargs={"project_pk": project1.pk, "pk": indicator_set.id})
    response = api_client1.put(
        url, data=_fs_put_payload(indicator_set, project1, fs), format="json"
    )
    assert response.status_code == 400
    assert expected_error_field in str(response.json())


@pytest.mark.parametrize(
    "override,coerced_field,expected_value",
    [
        # sector cleared for non-Business types
        (
            {
                "fs_type": "taf",
                "sector": "ce_waste_management",
                "number_of_solutions_supported_by": 1,
            },
            "sector",
            "",
        ),
        # geographical_coverage cleared for non-CTF types
        ({"fs_type": "business", "geographical_coverage": "national"}, "geographical_coverage", ""),
        # used_an_incubator cleared for types other than Business/Financial mechanism
        (
            {
                "fs_type": "taf",
                "used_an_incubator": "gfcr_funded",
                "number_of_solutions_supported_by": 1,
            },
            "used_an_incubator",
            None,
        ),
        # taf_name cleared for types other than Business/Financial mechanism
        (
            {"fs_type": "taf", "taf_name": "My TAF", "number_of_solutions_supported_by": 1},
            "taf_name",
            "",
        ),
        # taf_name cleared when used_an_incubator is not set (even for an allowed type)
        ({"fs_type": "business", "taf_name": "My TAF"}, "taf_name", ""),
        # local_enterprise cleared for types outside Financial facility/Business/Financial mechanism
        (
            {"fs_type": "taf", "local_enterprise": True, "number_of_solutions_supported_by": 1},
            "local_enterprise",
            False,
        ),
        # gender_smart cleared for types other than Business/Financial mechanism
        (
            {"fs_type": "taf", "gender_smart": True, "number_of_solutions_supported_by": 1},
            "gender_smart",
            False,
        ),
        # number_of_solutions_supported_by cleared for non-TAF types
        (
            {"fs_type": "business", "number_of_solutions_supported_by": 5},
            "number_of_solutions_supported_by",
            0,
        ),
        # sustainable_finance_mechanisms cleared for non-Financial mechanism types
        (
            {"fs_type": "business", "sustainable_finance_mechanisms": ["blue_bonds"]},
            "sustainable_finance_mechanisms",
            [],
        ),
    ],
)
def test_finance_solution_validate_coercion(
    db_setup,
    api_client1,
    project1,
    project_profile1,
    indicator_set,
    override,
    coerced_field,
    expected_value,
):
    project1.includes_gfcr = True
    project1.save()
    fs = {**_FS_BASE, **override}
    url = reverse("indicatorset-detail", kwargs={"project_pk": project1.pk, "pk": indicator_set.id})
    response = api_client1.put(
        url, data=_fs_put_payload(indicator_set, project1, fs), format="json"
    )
    assert (
        response.status_code == 200
    ), f"Expected 200, got {response.status_code}: {response.json()}"
    assert response.json()["finance_solutions"][0][coerced_field] == expected_value


def test_choices_new_gfcr_keys(db_setup, api_client1):
    url = reverse("choice-list")
    response = api_client1.get(url, format="json")
    assert response.status_code == 200
    choices_list = response.json()
    choices_by_name = {item["name"]: item["data"] for item in choices_list}

    assert "financesolutiontypes" in choices_by_name
    assert "geographicalcoverage" in choices_by_name
    assert "indicatorsettitles" in choices_by_name

    fs_type_ids = {c["id"] for c in choices_by_name["financesolutiontypes"]}
    assert "taf" in fs_type_ids
    assert "business" in fs_type_ids

    geo_ids = {c["id"] for c in choices_by_name["geographicalcoverage"]}
    assert "national" in geo_ids

    title_ids = {c["id"] for c in choices_by_name["indicatorsettitles"]}
    assert "Baseline" in title_ids
    assert "Phase 1 target" in title_ids


def test_report_new_columns_and_sheet_name(db_setup, project1):
    from openpyxl import load_workbook

    from api.reports.gfcr import create_report

    project1.includes_gfcr = True
    project1.save()

    indicator_set = GFCRIndicatorSet.objects.create(
        project=project1,
        title="Test IS",
        report_date="2024-02-10",
        indicator_set_type="report",
        f4_start_date=date(1970, 1, 1),
        f4_end_date=date(2024, 4, 19),
    )
    GFCRFinanceSolution.objects.create(
        indicator_set=indicator_set,
        name="Test FS",
        fs_type="ctf",
        geographical_coverage="national",
        taf_name="",
        number_of_solutions_supported_by=0,
    )

    report_path = create_report([project1.id])
    assert report_path is not None

    wb = load_workbook(report_path)
    assert "FacilitiesSolutions" in [ws.title for ws in wb.worksheets]
    assert "BusinessesFinanceSolutions" not in [ws.title for ws in wb.worksheets]

    fs_sheet = wb["FacilitiesSolutions"]
    row = list(fs_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    # type is at index 4, geographical_coverage at index 8
    assert row[4] == "Conservation trust fund (CTF)", f"Expected CTF display, got: {row}"
    assert row[8] == "national", f"Expected 'national', got: {row}"

    os.remove(report_path)
