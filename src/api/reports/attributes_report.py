from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT, Color, Font, PatternFill
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

from api.models import (
    SUPERUSER_APPROVED,
    BenthicAttribute,
    BenthicLifeHistory,
    FishFamily,
    FishGenus,
    FishGrouping,
    FishSpecies,
    Region,
)

FISH_FAMILY_NAME = "Fish Families"
FISH_GENERA_NAME = "Fish Genera"
FISH_SPECIES_NAME = "Fish Species"
FISH_GROUPINGS_NAME = "Fish Groupings"
BENTHIC_NAME = "Benthic"
YES = "Yes"
NO = "No"
YES_COLOR = Color(rgb="EAF7f0")
YES_COLOR_FILL = PatternFill(fgColor=YES_COLOR, fill_type="solid")
NO_COLOR = Color(rgb="FFFFFF")
NO_COLOR_FILL = PatternFill(fgColor=NO_COLOR, fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True)


def create_workbook_template():
    sheet_names = [
        FISH_FAMILY_NAME,
        FISH_GENERA_NAME,
        FISH_SPECIES_NAME,
        FISH_GROUPINGS_NAME,
        BENTHIC_NAME,
    ]
    wb = Workbook()
    DEFAULT_FONT.name = "Arial"

    wb.remove(wb.active)
    for sheet_name in sheet_names:
        wb.create_sheet(title=sheet_name)

    return wb


def insert_range_and_resize(ws, rng, data):
    _data = data or []
    row, col = coordinate_to_tuple(rng)
    dims = defaultdict(int)
    for n, row_val in enumerate(_data):
        for m, col_val in enumerate(row_val):
            if isinstance(col_val, tuple):
                val, fill = col_val
                ws.cell(row=row + n, column=col + m, value=val).fill = fill
            else:
                val = col_val
                ws.cell(row=row + n, column=col + m, value=val)

            # Bold header row
            if n == 0:
                ws.cell(row=row + n, column=col + m).font = HEADER_FONT

            width = len(str(val)) + 2
            dims[col + m] = max(dims[col + m], width)

    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value


def get_life_histories():
    return {str(lh.id): lh.name for lh in BenthicLifeHistory.objects.order_by("name")}


def get_regions():
    return {str(r.id): r.name for r in Region.objects.order_by("name")}


def create_m2m_row(lookups, data):
    if data is None:
        return []
    ids = [str(pk) for pk in data]
    return [(YES, YES_COLOR_FILL) if lookup in ids else (NO, NO_COLOR_FILL) for lookup in lookups]


def write_fish_families(wb, regions):
    region_names = list(regions.values())
    COLUMN_NAMES = [
        "Name",
        "Biomass Constant A",
        "Biomass Constant B",
        "Biomass Constant C",
    ] + region_names

    data = [
        COLUMN_NAMES,
        *[
            [
                fish_family.name,
                fish_family.biomass_constant_a,
                fish_family.biomass_constant_b,
                fish_family.biomass_constant_c,
                *create_m2m_row(
                    regions, [str(r) for r in fish_family.regions] if fish_family.regions else None
                ),
            ]
            for fish_family in FishFamily.objects.filter(status=SUPERUSER_APPROVED).order_by("name")
        ],
    ]

    insert_range_and_resize(wb[FISH_FAMILY_NAME], "A1", data)


def write_fish_genera(wb, regions):
    region_names = list(regions.values())
    COLUMN_NAMES = [
        "Family",
        "Name",
        "Biomass Constant A",
        "Biomass Constant B",
        "Biomass Constant C",
    ] + region_names

    data = [
        COLUMN_NAMES,
        *[
            [
                fish_genus.family.name,
                fish_genus.name,
                fish_genus.biomass_constant_a,
                fish_genus.biomass_constant_b,
                fish_genus.biomass_constant_c,
                *create_m2m_row(
                    regions, [str(r) for r in fish_genus.regions] if fish_genus.regions else None
                ),
            ]
            for fish_genus in FishGenus.objects.select_related("family")
            .filter(status=SUPERUSER_APPROVED)
            .order_by("family__name", "name")
        ],
    ]
    insert_range_and_resize(wb[FISH_GENERA_NAME], "A1", data)


def write_fish_species(wb, regions):
    region_names = list(regions.values())
    COLUMN_NAMES = [
        "Family",
        "Name",
        "Biomass Constant A",
        "Biomass Constant B",
        "Biomass Constant C",
        "Trophic Group",
        "Functional Group",
        "Max Length (cm)",
        "Max Type",
        "Group Size",
        "Trophic Level",
        "Vulnerability",
        "Climate Score",
    ] + region_names

    data = [
        COLUMN_NAMES,
        *[
            [
                fish_species.genus.family.name,
                str(fish_species),
                fish_species.biomass_constant_a,
                fish_species.biomass_constant_b,
                fish_species.biomass_constant_c,
                fish_species.trophic_group and fish_species.trophic_group.name,
                fish_species.functional_group and fish_species.functional_group.name,
                fish_species.max_length,
                fish_species.max_length_type,
                fish_species.group_size and fish_species.group_size.name,
                fish_species.trophic_level,
                fish_species.vulnerability,
                fish_species.climate_score,
                *create_m2m_row(regions, fish_species.regions.values_list("id", flat=True)),
            ]
            for fish_species in FishSpecies.objects.select_related(
                "genus",
                "genus__family",
                "trophic_group",
                "functional_group",
                "group_size",
            )
            .prefetch_related("regions")
            .filter(status=SUPERUSER_APPROVED)
            .order_by("genus__family__name", "name")
        ],
    ]
    insert_range_and_resize(wb[FISH_SPECIES_NAME], "A1", data)


def write_fish_grouping(wb, regions):
    region_names = list(regions.values())
    COLUMN_NAMES = [
        "Name",
        "Fish Taxa",
        "Biomass Constant A",
        "Biomass Constant B",
        "Biomass Constant C",
    ] + region_names

    data = [
        COLUMN_NAMES,
        *[
            [
                fish_group.name,
                ",".join([str(a.attribute) for a in fish_group.attribute_grouping.all()]),
                fish_group.biomass_constant_a,
                fish_group.biomass_constant_b,
                fish_group.biomass_constant_c,
                *create_m2m_row(regions, fish_group.regions.values_list("id", flat=True)),
            ]
            for fish_group in FishGrouping.objects.select_related()
            .filter(status=SUPERUSER_APPROVED)
            .order_by("name")
        ],
    ]
    insert_range_and_resize(wb[FISH_GROUPINGS_NAME], "A1", data)


def write_benthic(wb, regions):
    life_histories = get_life_histories()
    life_history_names = list(life_histories.values())
    region_names = list(regions.values())
    COLUMN_NAMES = (
        [
            "Name",
            "Parent",
            "Category",
        ]
        + life_history_names
        + region_names
    )

    data = [
        COLUMN_NAMES,
        *[
            [
                ba.name,
                ba.parent and ba.parent.name,
                ba.origin and ba.origin.name,
                *create_m2m_row(life_histories, ba.life_histories.values_list("id", flat=True)),
                *create_m2m_row(regions, ba.regions.values_list("id", flat=True)),
            ]
            for ba in BenthicAttribute.objects.filter(status=SUPERUSER_APPROVED).order_by("name")
        ],
    ]
    insert_range_and_resize(wb[BENTHIC_NAME], "A1", data)


def write_attribute_reference(output_path):
    wb = create_workbook_template()
    regions = get_regions()
    write_fish_families(wb, regions)
    write_fish_genera(wb, regions)
    write_fish_species(wb, regions)
    write_fish_grouping(wb, regions)
    write_benthic(wb, regions)

    wb.save(output_path)
