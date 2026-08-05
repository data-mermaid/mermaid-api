import os
from collections.abc import Iterable
from typing import Any

from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..utils.castutils import cast_str_value


def _get_xlsx_template(template: str):
    if not template.endswith(".xlsx"):
        template = f"{template}.xlsx"

    for report_tmpl_dir in settings.REPORT_TEMPLATES:
        tpl_path = os.path.join(report_tmpl_dir, template)
        if os.path.exists(tpl_path):
            return tpl_path

    raise ValueError(f"Template [{template}] not found.")


def get_workbook(workbook: Workbook | str | None = None) -> Workbook:
    if isinstance(workbook, str):
        tpl = _get_xlsx_template(workbook)
        wb = load_workbook(tpl)
    elif isinstance(workbook, Workbook):
        wb = workbook
    else:
        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)

    return wb


def get_worksheet(wb: Workbook, name: str | None = None, create: bool = False) -> Worksheet:
    if name:
        name = name.lower()
        for ws in wb.worksheets:
            if ws.title.lower() == name:
                return ws

        if create:
            return wb.create_sheet(title=name)

        raise ValueError(f"Worksheet [{name}] not found.")

    else:
        return wb.worksheets[0]


def write_data_to_sheet(
    workbook: Workbook | str | None,
    sheet_name: str,
    data: Iterable[Iterable[Any]],
    row: int = 1,
    col: int = 1,
) -> tuple[int, int]:
    wb = get_workbook(workbook)
    ws = get_worksheet(wb, sheet_name, create=True)

    current_row = row
    current_col = col
    for n, row_val in enumerate(data):
        current_row = row + n
        for m, col_val in enumerate(row_val):
            current_col = col + m
            ws.cell(row=current_row, column=current_col).value = cast_str_value(col_val)

    return current_row, current_col


def auto_size_columns(worksheet, max_width=60):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                print(f"Auto sizing excel columns: {e}")
                pass

        adjusted_width = min(max_length + 2, max_width)
        worksheet.column_dimensions[column].width = adjusted_width
