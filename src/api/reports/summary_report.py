import csv
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT, Font
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

from ..mocks import MockRequest
from ..resources.sampleunitmethods.beltfishmethod import (
    BeltFishProjectMethodObsView,
    BeltFishProjectMethodSUView,
    BeltFishProjectMethodSEView,
)
from ..resources.sampleunitmethods.benthicpitmethod import (
    BenthicPITProjectMethodObsView,
    BenthicPITProjectMethodSUView,
    BenthicPITProjectMethodSEView,
)
from ..resources.sampleunitmethods.benthiclitmethod import (
    BenthicLITProjectMethodObsView,
    BenthicLITProjectMethodSUView,
    BenthicLITProjectMethodSEView,
)
from ..resources.sampleunitmethods.habitatcomplexitymethod import (
    HabitatComplexityProjectMethodObsView,
    HabitatComplexityProjectMethodSUView,
    HabitatComplexityProjectMethodSEView,
)
from ..resources.sampleunitmethods.bleachingquadratcollectionmethod import (
    BleachingQCProjectMethodObsColoniesBleachedView,
    BleachingQCProjectMethodObsQuadratBenthicPercentView,
    BleachingQCProjectMethodSUView,
    BleachingQCProjectMethodSEView,
)
from ..resources.sampleunitmethods.benthicphotoquadrattransectmethod import (
    BenthicPQTProjectMethodObsView,
    BenthicPQTProjectMethodSUView,
    BenthicPQTProjectMethodSEView,
)

DEFAULT_FONT_NAME = "Arial"
HEADER_FONT = Font(name=DEFAULT_FONT_NAME, bold=True)


def insert_range_and_resize(ws, rng, data, bold_header=True):
    _data = data or []
    row, col = coordinate_to_tuple(rng)
    dims = defaultdict(int)
    for n, row_val in enumerate(_data):
        for m, val in enumerate(row_val):
            ws.cell(row=row + n, column=col + m, value=val)

            # Bold header row
            if n == 0 and bold_header:
                ws.cell(row=row + n, column=col + m).font = HEADER_FONT

            width = len(str(val)) + 2
            dims[col + m] = max(dims[col + m], width)

    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value


def get_viewset_csv_content(view_cls, project_pk, request):
    request = MockRequest(query_params={"field_report": "t"})
    vw = view_cls()
    vw.kwargs = {"project_pk": project_pk}
    vw.request = request
    resp = vw.csv(request, **vw.kwargs)
    return list(
        csv.reader([str(row, "UTF-8").strip() for row in resp.streaming_content])
    )


def create_workbook_template(sheet_names):
    wb = Workbook()
    DEFAULT_FONT.name = DEFAULT_FONT_NAME

    wb.remove(wb.active)
    for sheet_name in sheet_names:
        wb.create_sheet(title=sheet_name)

    return wb


def create_belt_fish_report(request, project_pk):
    worksheet_names = [
        "Belt Fish Obs",
        "Belt Fish SU",
        "Belt Fish SE",
    ]
    wb = create_workbook_template(worksheet_names)

    content = get_viewset_csv_content(BeltFishProjectMethodObsView, project_pk, request)
    insert_range_and_resize(wb[worksheet_names[0]], "A1", content)

    content = get_viewset_csv_content(BeltFishProjectMethodSUView, project_pk, request)
    insert_range_and_resize(wb[worksheet_names[1]], "A1", content)

    content = get_viewset_csv_content(BeltFishProjectMethodSEView, project_pk, request)
    insert_range_and_resize(wb[worksheet_names[2]], "A1", content)

    return wb


def create_benthic_pit_report(request, project_pk):
    worksheet_names = [
        "Benthic PIT Obs",
        "Benthic PIT SU",
        "Benthic PIT SE",
    ]
    wb = create_workbook_template(worksheet_names)

    content = get_viewset_csv_content(
        BenthicPITProjectMethodObsView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[0]], "A1", content)

    content = get_viewset_csv_content(
        BenthicPITProjectMethodSUView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[1]], "A1", content)

    content = get_viewset_csv_content(
        BenthicPITProjectMethodSEView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[2]], "A1", content)

    return wb


def create_benthic_lit_report(request, project_pk):
    worksheet_names = [
        "Benthic LIT Obs",
        "Benthic LIT SU",
        "Benthic LIT SE",
    ]
    wb = create_workbook_template(worksheet_names)

    content = get_viewset_csv_content(
        BenthicLITProjectMethodObsView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[0]], "A1", content)

    content = get_viewset_csv_content(
        BenthicLITProjectMethodSUView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[1]], "A1", content)

    content = get_viewset_csv_content(
        BenthicLITProjectMethodSEView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[2]], "A1", content)

    return wb


def create_bleaching_qc_report(request, project_pk):
    worksheet_names = [
        "BQT Colonies Bleached Obs",
        "BQT Quad Benthic Percent Obs",
        "Bleaching QT SU",
        "Bleaching QT SE",
    ]

    wb = create_workbook_template(worksheet_names)

    content = get_viewset_csv_content(
        BleachingQCProjectMethodObsColoniesBleachedView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[0]], "A1", content)

    content = get_viewset_csv_content(
        BleachingQCProjectMethodObsQuadratBenthicPercentView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[1]], "A1", content)

    content = get_viewset_csv_content(
        BleachingQCProjectMethodSUView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[2]], "A1", content)

    content = get_viewset_csv_content(
        BleachingQCProjectMethodSEView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[3]], "A1", content)

    return wb


def create_benthic_pqt_report(request, project_pk):
    worksheet_names = [
        "Benthic PQT Obs",
        "Benthic PQT SU",
        "Benthic PQT SE",
    ]
    wb = create_workbook_template(worksheet_names)

    content = get_viewset_csv_content(
        BenthicPQTProjectMethodObsView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[0]], "A1", content)

    content = get_viewset_csv_content(
        BenthicPQTProjectMethodSUView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[1]], "A1", content)

    content = get_viewset_csv_content(
        BenthicPQTProjectMethodSEView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[2]], "A1", content)

    return wb


def create_habitat_complexity_report(request, project_pk):
    worksheet_names = [
        "Habitat Complexity Obs",
        "Habitat Complexity SU",
        "Habitat Complexity SE",
    ]
    wb = create_workbook_template(worksheet_names)

    content = get_viewset_csv_content(
        HabitatComplexityProjectMethodObsView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[0]], "A1", content)

    content = get_viewset_csv_content(
        HabitatComplexityProjectMethodSUView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[1]], "A1", content)

    content = get_viewset_csv_content(
        HabitatComplexityProjectMethodSEView, project_pk, request
    )
    insert_range_and_resize(wb[worksheet_names[2]], "A1", content)

    return wb
